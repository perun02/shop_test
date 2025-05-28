from datetime import datetime
import os
from aiogram import Router, F
from aiogram.filters import Command
from aiogram.types import Message, FSInputFile
import openpyxl
from bot.models import Order, OrderItem

router = Router()

# команда для выгрузки всех заказов в Excel файл
# создает таблицу с информацией о заказах, включая данные покупателя и товары
# автоматически настраивает ширину столбцов и отправляет файл администратору
@router.message(Command("admin_xlsx"))
async def handle_admin_xlsx(message: Message):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заказы"

    # Заголовки
    headers = [
        "ID заказа", 
        "Дата заказа",
        "ФИО получателя",
        "Телефон",
        "Адрес",
        "Статус оплаты",
        "Telegram ID",
        "Username",
        "Товары"
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    row = 2
    async for order in Order.objects.select_related('user').prefetch_related('items__product__subcategory').all():
        items_info = []
        async for item in order.items.all():
            items_info.append(f"{item.product.subcategory.name} x {item.quantity}")
        items_str = ", ".join(items_info)

        row_data = [
            order.id,
            order.created_at.strftime("%d.%m.%Y %H:%M"),
            order.full_name,
            order.phone_number,
            order.address,
            "Оплачен" if order.is_paid else "Не оплачен",
            order.user.telegram_id,
            order.user.username or "Нет username",
            items_str
        ]
        
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    for column in ws.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    filename = f"orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = f"media/{filename}"
    wb.save(filepath)
    document = FSInputFile(filepath)
    await message.answer_document(
        document=document,
        caption="Отчет по всем заказам"
    )

    os.remove(filepath)
